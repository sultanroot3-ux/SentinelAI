"""Visitor reports: CSV / JSON / PDF / Excel."""
import csv
import io
from datetime import date, datetime, time, timedelta

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response, StreamingResponse
from sqlalchemy.orm import Session, joinedload

from app.api.utils import serialize_log
from app.core.security import require_roles
from app.db.database import get_db
from app.models.models import RecognitionLog, User

router = APIRouter(prefix="/api/reports", tags=["reports"])

_HEADERS = ["ID", "User ID", "User Name", "Camera", "Score", "Snapshot", "Timestamp"]


def _log_row(log: RecognitionLog) -> list:
    return [
        log.id,
        log.user_id if log.user_id is not None else "",
        log.user.name if log.user else "",
        log.camera,
        log.score if log.score is not None else "",
        log.snapshot_url or "",
        log.timestamp.isoformat(sep=" ", timespec="seconds"),
    ]


def _build_pdf(logs, period: str, start: datetime) -> bytes:
    """Tabular PDF via reportlab platypus."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=15 * mm, rightMargin=15 * mm, topMargin=15 * mm, bottomMargin=15 * mm,
        title=f"SentinelAI visitor report ({period})",
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph("SentinelAI — Visitor Report", styles["Title"]),
        Paragraph(
            f"Period: {period} (from {start.date().isoformat()} UTC) · "
            f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')} · "
            f"Total recognitions: {len(logs)}",
            styles["Normal"],
        ),
        Spacer(1, 6 * mm),
    ]
    data = [_HEADERS] + [[str(c) for c in _log_row(l)] for l in logs]
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#131a26")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#94a3b8")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(table)
    doc.build(story)
    return buf.getvalue()


def _build_xlsx(logs, period: str, start: datetime) -> bytes:
    """Excel workbook via openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = f"Visitors ({period})"
    ws.append(_HEADERS)
    header_fill = PatternFill("solid", fgColor="131A26")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for log in logs:
        ws.append(_log_row(log))
    for col in range(1, len(_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _period_start(period: str) -> datetime:
    # Timestamps are stored in UTC, so the report window must be UTC too
    # (date.today() is local time and drifts near midnight).
    today = datetime.combine(datetime.utcnow().date(), time.min)
    if period == "daily":
        return today
    if period == "weekly":
        return today - timedelta(days=6)
    # monthly
    return today - timedelta(days=29)


@router.get("/visitors")
def visitors_report(
    period: str = Query("daily", pattern="^(daily|weekly|monthly)$"),
    format: str = Query("csv", pattern="^(csv|json|pdf|xlsx)$"),
    db: Session = Depends(get_db),
    _: User = Depends(require_roles("admin", "security_officer")),
):
    start = _period_start(period)
    logs = (
        db.query(RecognitionLog)
        .options(joinedload(RecognitionLog.user))
        .filter(RecognitionLog.timestamp >= start)
        .order_by(RecognitionLog.timestamp.desc())
        .all()
    )
    stamp = f"visitors_{period}_{date.today().isoformat()}"

    if format == "pdf":
        try:
            content = _build_pdf(logs, period, start)
        except ImportError:
            raise HTTPException(503, detail="PDF export requires the 'reportlab' package.")
        return Response(
            content,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{stamp}.pdf"'},
        )

    if format == "xlsx":
        try:
            content = _build_xlsx(logs, period, start)
        except ImportError:
            raise HTTPException(503, detail="Excel export requires the 'openpyxl' package.")
        return Response(
            content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{stamp}.xlsx"'},
        )

    if format == "json":
        return {
            "period": period,
            "date_from": start.isoformat(),
            "generated_at": datetime.utcnow().isoformat(),
            "total": len(logs),
            "items": [serialize_log(l).model_dump(mode="json") for l in logs],
        }

    def csv_rows():
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(
            ["id", "user_id", "user_name", "camera", "score", "snapshot_url", "timestamp"]
        )
        yield buffer.getvalue()
        for log in logs:
            buffer.seek(0)
            buffer.truncate(0)
            writer.writerow(_log_row(log))
            yield buffer.getvalue()

    return StreamingResponse(
        csv_rows(),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{stamp}.csv"'},
    )
